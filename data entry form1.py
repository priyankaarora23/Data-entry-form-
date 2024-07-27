from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
import pathlib
from openpyxl import Workbook
import openpyxl
from tkinter import messagebox  # Ensure messagebox is imported

root = Tk()
root.title("Data Enter")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file = pathlib.Path('Backend_data.xlsx')

if not file.exists():
    wb = Workbook()  # Create a new workbook
    sheet = wb.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    wb.save('Backend_data.xlsx')

def submit():
    wb = openpyxl.load_workbook('Backend_data.xlsx')
    sheet = wb.active
    
    row = sheet.max_row + 1
    sheet[f'A{row}'] = nameValue.get()
    sheet[f'B{row}'] = ContactValue.get()
    sheet[f'C{row}'] = AgeValue.get()
    sheet[f'D{row}'] = gender_combobox.get()
    sheet[f'E{row}'] = addressEntry.get("1.0", tk.END).strip()
    
    wb.save('Backend_data.xlsx')
    messagebox.showinfo("Success", "Data saved successfully!")  # Show success message
    clear()

def clear():
    nameValue.set('')
    ContactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)

# Heading
Label(root, text="Please fill out this entry form:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Contact No.', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=370, y=200)
Label(root, text='Address', font=23, bg="#326273", fg="#fff").place(x=50, y=250)

# Entry fields
nameValue = StringVar()
ContactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
ContactEntry = Entry(root, textvariable=ContactValue, width=45, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)

# Gender Combobox
gender_combobox = Combobox(root, values=['Male', 'Female'], font='arial 14', state='readonly', width=14)
gender_combobox.place(x=440, y=200)
gender_combobox.set('Male')
addressEntry = Text(root, width=50, height=4, bd=4)

# Place Entry fields
nameEntry.place(x=200, y=100)
ContactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)
addressEntry.place(x=200, y=250)

Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=lambda: root.destroy()).place(x=480, y=350)

root.mainloop()
